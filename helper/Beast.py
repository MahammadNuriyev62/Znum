import znum.Znum as xusun
from uuid import uuid4


class Beast:
    ZNUM_SIZE = 8

    class Methods:
        TOPSIS = 1
        PROMETHEE = 2

    @staticmethod
    def read_xlsx_main():
        filename = Beast.get_file_path()
        return Beast.read_xlsx(filename)

    @staticmethod
    def read_znums_from_xlsx(method=None):
        method = method or Beast.Methods.TOPSIS
        table = Beast.read_xlsx_main()
        if method == Beast.Methods.TOPSIS:
            return Beast.parse_znums_from_table(table)
        elif method == Beast.Methods.PROMETHEE:
            return Beast.parse_znums_from_table(table)
        else:
            raise Exception("Invalid Optimization Method for input Table")

    @staticmethod
    def znum_to_row(znum):
        """
        :type znum: xusun.Znum
        """
        return [*znum.A, *znum.B]

    @staticmethod
    def transpose(matrix):
        return zip(*matrix)

    @staticmethod
    def save_znums_to_xlsx(data):
        """
        :type data: list[xusun.Znum] or list[list[xusun.Znum]]
        """
        data_new = []
        for d in data:
            if type(d) == xusun.Znum:
                data_new.append(Beast.znum_to_row(d))
            else:
                row = []
                for dd in d:
                    row.extend(Beast.znum_to_row(dd))
                data_new.append(row)
        Beast.save_array_in_excel(data_new)
        # method = method or Beast.Methods.TOPSIS
        # table = Beast.read_xlsx_main()
        # if method == Beast.Methods.TOPSIS:
        #     return Beast.parse_znums_from_table(table)
        # elif method == Beast.Methods.PROMETHEE:
        #     return Beast.parse_znums_from_table(table)
        # else:
        #     raise Exception('Invalid Optimization Method for input Table')

    @staticmethod
    def get_file_path():
        import tkinter.filedialog

        filename = tkinter.filedialog.askopenfile()
        return filename.name

    @staticmethod
    def read_xlsx(filename: str):
        from openpyxl import load_workbook, Workbook

        workbook: Workbook = load_workbook(filename)

        table = []
        for sheet in workbook:
            for row in sheet.rows:
                values = [o.value for o in row]
                any(values) and table.append(values)

        return table

    @staticmethod
    def validate_topsis_table(weights, alternatives, types):
        if weights[0] != "W":
            raise Exception("First row must be weights")
        if types[0] != "T":
            raise Exception("Last row must be types")
        filtered_weights = [w for w in weights[1:] if w]
        if len(filtered_weights) != len(alternatives[0]) - 1:
            raise Exception("Weights and Alternatives must have same number of columns")
        filtered_types = [t for t in types[1:] if t]
        if len(filtered_types) != len(filtered_weights) // 8:
            raise Exception(
                "There must be the same number of types as the number weights"
            )
        return True

    @staticmethod
    def parse_znums_from_table(table: list[list]):
        Beast.validate_topsis_table(table[0], table[1:-1], table[-1])
        weights, alternatives, types = table[0], table[1:-1], table[-1]
        weights_modified = Beast.parse_znums_from_row(weights[1:])
        alternatives_modified = [
            Beast.parse_znums_from_row(row[1:]) for row in alternatives
        ]
        types_modified = [t for t in types[1:] if t]
        return [weights_modified, *alternatives_modified, types_modified]

    @staticmethod
    def parse_znums_from_row(row: list[int]):
        row_modified = []
        for i in range(0, len(row), Beast.ZNUM_SIZE):
            znumAsList = row[i : i + Beast.ZNUM_SIZE]
            index = int(Beast.ZNUM_SIZE / 2)
            znum = xusun.Znum(A=znumAsList[:index], B=znumAsList[index:])
            row_modified.append(znum)
        return row_modified

    @staticmethod
    def save_znums_as_one_column_grouped_by_criteria(table):
        from openpyxl import Workbook

        table: list[list[xusun.Znum]]

        workbook = Workbook()
        sheet = workbook.create_sheet("XUSUN")
        table_transpose = zip(*table)
        for row in table_transpose:
            for znum in row:
                sheet.append(znum.A + znum.B)
        workbook.save("output.xlsx")

    @staticmethod
    def timer(func):
        import time

        def wrapper():
            start = time.time()
            result = func()
            end = time.time()
            print(f"execution time : {end - start}")
            return result

        return wrapper

    @staticmethod
    def save_array_in_excel(*arrays):
        if type(arrays[-1]) == str:
            *arrays, filename = arrays
        else:
            filename = f"output_{uuid4()}.xlsx"

        print(f"{filename = }")

        spacing = 3
        from openpyxl import Workbook

        workbook = Workbook()

        sheet = workbook.active
        for array in arrays:
            if type(array[0]) is list or type(array[0]) is tuple:
                for row in array:
                    sheet.append(row)
            else:
                sheet.append(array)

            for i in range(spacing):
                sheet.append([None])

        workbook.save(filename)
